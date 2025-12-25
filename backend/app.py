#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ZVD Area Calculator - Простой калькулятор площадей разверток
"""

from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
import logging
from pathlib import Path
import os

# Версия приложения
try:
    version_file = Path(__file__).parent.parent / 'VERSION'
    if version_file.exists():
        with open(version_file, 'r', encoding='utf-8') as f:
            VERSION = f.read().strip()
    else:
        VERSION = "1.0.0"
except:
    VERSION = "1.0.0"

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('backend.log', encoding='utf-8'),  # Логи в файл
        logging.StreamHandler()  # Логи в консоль
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)

# Папка для загрузки файлов
UPLOAD_FOLDER = Path('uploads')
UPLOAD_FOLDER.mkdir(exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50 MB max для больших DXF файлов
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

from utils.dxf_parser import parse_dxf_dimensions
from utils.area_calculator import calculate_total_area

# Импорты для работы с Excel и PDF
try:
    import pandas as pd
    import openpyxl
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    logger.warning("pandas/openpyxl не установлены, импорт Excel недоступен")

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False
    logger.warning("pdfplumber не установлен, импорт PDF недоступен")

# Импорты для экспорта в PDF
try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether
    from reportlab.graphics.shapes import Drawing, Rect, Group
    from reportlab.graphics import renderPDF
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("reportlab не установлен, экспорт PDF недоступен")

# Импортируем blueprint для раскроя
try:
    logger.info("Попытка импорта API раскроя...")
    from api.nesting import nesting_bp
    logger.info("[OK] API раскроя успешно импортирован")
    app.register_blueprint(nesting_bp, url_prefix='/api/nesting')
    logger.info("[OK] Blueprint раскроя зарегистрирован: /api/nesting")
except ImportError as e:
    logger.error(f"[ERROR] Не удалось загрузить API раскроя: {e}", exc_info=True)
except Exception as e:
    logger.error(f"[ERROR] Ошибка при регистрации blueprint раскроя: {e}", exc_info=True)

@app.route('/')
def index():
    return jsonify({
        'name': 'ZVD Area Calculator API',
        'version': VERSION,
        'description': 'Простой калькулятор площадей разверток из DXF',
        'endpoints': {
            '/api/upload': 'POST - Загрузить DXF файлы',
            '/api/calculate': 'POST - Рассчитать площади',
            '/api/import/excel': 'POST - Импорт данных из Excel',
            '/api/import/pdf': 'POST - Импорт данных из PDF',
            '/api/nesting/validate': 'POST - Валидация раскроя (проверка координат)'
        }
    })

@app.route('/api/version', methods=['GET'])
def get_version():
    """Возвращает версию приложения"""
    return jsonify({'version': VERSION, 'name': 'ZVD Nesting Calculator'})

@app.route('/api/upload', methods=['POST'])
def upload_files():
    """
    Загрузка DXF файлов
    
    POST /api/upload
    Files: multiple DXF files
    
    Returns:
        [
            {'name': 'деталь.dxf', 'width': 1500, 'height': 400, 'area_m2': 0.6},
            ...
        ]
    """
    try:
        logger.info(f"[UPLOAD] Получен запрос на загрузку файлов")
        
        if 'files' not in request.files:
            logger.warning("[UPLOAD] Нет файлов в запросе")
            return jsonify({'error': 'No files provided'}), 400
        
        files = request.files.getlist('files')
        logger.info(f"[UPLOAD] Получено файлов: {len(files)}")
        results = []
        
        for idx, file in enumerate(files):
            logger.info(f"[UPLOAD] Обработка файла {idx+1}/{len(files)}: {file.filename}")
            if file.filename == '':
                continue
            
            if not file.filename.lower().endswith('.dxf'):
                continue
            
            # Сохраняем оригинальное имя файла для отображения
            original_filename = file.filename
            
            # Для сохранения на диск используем безопасное имя (чтобы избежать проблем с файловой системой)
            # Но для отображения используем оригинальное имя
            safe_filename = secure_filename(original_filename)
            # Если secure_filename создал пустое имя, используем временное
            if not safe_filename:
                import uuid
                safe_filename = f"file_{uuid.uuid4().hex[:8]}.dxf"
            
            filepath = UPLOAD_FOLDER / safe_filename
            logger.info(f"[UPLOAD] Сохранение файла: {filepath}")
            file.save(filepath)
            logger.info(f"[UPLOAD] Файл сохранен, размер: {filepath.stat().st_size} байт")
            
            # Парсим размеры
            try:
                logger.info(f"[UPLOAD] Начало парсинга DXF: {filepath}")
                width, height = parse_dxf_dimensions(str(filepath))
                logger.info(f"[UPLOAD] Парсинг завершен: {width}x{height}")
                
                if width and height:
                    area_m2 = (width * height) / 1_000_000
                    
                    # Используем ОРИГИНАЛЬНОЕ имя файла для отображения
                    results.append({
                        'name': original_filename,  # Оригинальное имя с русскими буквами и пробелами
                        'width': round(width, 1),
                        'height': round(height, 1),
                        'area_m2': round(area_m2, 4),
                        'quantity': 1  # По умолчанию 1
                    })
                    
                    logger.info(f"✓ {original_filename}: {width:.0f}×{height:.0f} мм")
                else:
                    logger.warning(f"✗ {original_filename}: не удалось определить размеры (width={width}, height={height})")
                    # Добавляем файл даже если не удалось определить размеры (с нулевыми размерами)
                    results.append({
                        'name': original_filename,
                        'width': 0,
                        'height': 0,
                        'area_m2': 0,
                        'quantity': 1
                    })
            except Exception as parse_error:
                logger.error(f"✗ {original_filename}: ошибка парсинга DXF: {parse_error}", exc_info=True)
                # Добавляем файл даже при ошибке парсинга
                results.append({
                    'name': original_filename,
                    'width': 0,
                    'height': 0,
                    'area_m2': 0,
                    'quantity': 1
                })
        
        logger.info(f"[UPLOAD] Успешно обработано файлов: {len(results)}")
        return jsonify({
            'success': True,
            'parts': results
        })
        
    except Exception as e:
        logger.error(f"[UPLOAD] Ошибка загрузки файлов: {e}", exc_info=True)
        import traceback
        logger.error(f"[UPLOAD] Traceback: {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/calculate', methods=['POST'])
def calculate():
    """
    Расчет итоговой площади
    
    POST /api/calculate
    {
        "parts": [
            {"name": "деталь.dxf", "width": 1500, "height": 400, "quantity": 2},
            ...
        ],
        "cut_gap": 5,
        "edge_margin": 10
    }
    
    Returns:
        {
            "total_area_m2": 1.23,
            "total_area_with_gaps_m2": 1.35,
            "parts_count": 5
        }
    """
    try:
        data = request.get_json()
        
        parts = data.get('parts', [])
        cut_gap = data.get('cut_gap', 5)
        edge_margin = data.get('edge_margin', 10)
        
        result = calculate_total_area(parts, cut_gap, edge_margin)
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Ошибка расчета: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/import/excel', methods=['POST'])
def import_excel():
    """
    Импорт данных из Excel файла
    
    POST /api/import/excel
    File: Excel file (.xlsx, .xls)
    
    Ожидаемый формат Excel:
    - Колонки: Название, Ширина (мм), Высота (мм), Количество (опционально)
    - Первая строка - заголовки
    
    Returns:
        {
            'success': True,
            'parts': [
                {'name': 'деталь', 'width': 1500, 'height': 400, 'area_m2': 0.6, 'quantity': 1},
                ...
            ]
        }
    """
    try:
        if not PANDAS_AVAILABLE:
            return jsonify({'error': 'pandas/openpyxl не установлены. Установите: pip install pandas openpyxl'}), 500
        
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not (file.filename.lower().endswith('.xlsx') or file.filename.lower().endswith('.xls')):
            return jsonify({'error': 'Файл должен быть Excel (.xlsx или .xls)'}), 400
        
        # Сохраняем файл временно
        safe_filename = secure_filename(file.filename)
        if not safe_filename:
            import uuid
            safe_filename = f"excel_{uuid.uuid4().hex[:8]}.xlsx"
        
        filepath = UPLOAD_FOLDER / safe_filename
        file.save(filepath)
        
        # Читаем Excel
        try:
            df = pd.read_excel(filepath, engine='openpyxl' if file.filename.lower().endswith('.xlsx') else None)
            
            # Ищем колонки (пробуем разные варианты названий)
            name_col = None
            width_col = None
            height_col = None
            quantity_col = None
            
            for col in df.columns:
                col_lower = str(col).lower()
                if 'название' in col_lower or 'имя' in col_lower or 'name' in col_lower or 'деталь' in col_lower:
                    name_col = col
                elif 'ширина' in col_lower or 'width' in col_lower:
                    width_col = col
                elif 'высота' in col_lower or 'height' in col_lower:
                    height_col = col
                elif 'количество' in col_lower or 'quantity' in col_lower or 'кол-во' in col_lower:
                    quantity_col = col
            
            if not name_col or not width_col or not height_col:
                return jsonify({
                    'error': 'Не найдены необходимые колонки. Ожидаются: Название, Ширина (мм), Высота (мм)'
                }), 400
            
            # Парсим данные
            results = []
            for idx, row in df.iterrows():
                try:
                    name = str(row[name_col]).strip() if pd.notna(row[name_col]) else f'Деталь {idx + 1}'
                    width = float(row[width_col]) if pd.notna(row[width_col]) else 0
                    height = float(row[height_col]) if pd.notna(row[height_col]) else 0
                    quantity = int(row[quantity_col]) if quantity_col and pd.notna(row[quantity_col]) else 1
                    
                    if width > 0 and height > 0:
                        area_m2 = (width * height) / 1_000_000
                        results.append({
                            'name': name,
                            'width': round(width, 1),
                            'height': round(height, 1),
                            'area_m2': round(area_m2, 4),
                            'quantity': quantity
                        })
                except Exception as row_error:
                    logger.warning(f"Ошибка обработки строки {idx + 1}: {row_error}")
                    continue
            
            # Удаляем временный файл
            try:
                filepath.unlink()
            except:
                pass
            
            logger.info(f"✓ Импортировано {len(results)} деталей из Excel")
            
            return jsonify({
                'success': True,
                'parts': results
            })
            
        except Exception as parse_error:
            logger.error(f"Ошибка парсинга Excel: {parse_error}", exc_info=True)
            return jsonify({'error': f'Ошибка чтения Excel: {str(parse_error)}'}), 500
        
    except Exception as e:
        logger.error(f"Ошибка импорта Excel: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/import/pdf', methods=['POST'])
def import_pdf():
    """
    Импорт данных из PDF файла
    
    POST /api/import/pdf
    File: PDF file (.pdf)
    
    Пытается извлечь таблицы из PDF и найти данные о деталях
    
    Returns:
        {
            'success': True,
            'parts': [
                {'name': 'деталь', 'width': 1500, 'height': 400, 'area_m2': 0.6, 'quantity': 1},
                ...
            ]
        }
    """
    try:
        if not PDFPLUMBER_AVAILABLE:
            return jsonify({'error': 'pdfplumber не установлен. Установите: pip install pdfplumber'}), 500
        
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({'error': 'Файл должен быть PDF (.pdf)'}), 400
        
        # Сохраняем файл временно
        safe_filename = secure_filename(file.filename)
        if not safe_filename:
            import uuid
            safe_filename = f"pdf_{uuid.uuid4().hex[:8]}.pdf"
        
        filepath = UPLOAD_FOLDER / safe_filename
        file.save(filepath)
        
        # Читаем PDF
        try:
            results = []
            
            with pdfplumber.open(filepath) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    # Пытаемся извлечь таблицы
                    tables = page.extract_tables()
                    
                    for table in tables:
                        if not table or len(table) < 2:
                            continue
                        
                        # Первая строка - заголовки
                        headers = [str(cell).lower() if cell else '' for cell in table[0]]
                        
                        # Ищем индексы колонок
                        name_idx = None
                        width_idx = None
                        height_idx = None
                        quantity_idx = None
                        
                        for i, header in enumerate(headers):
                            if 'название' in header or 'имя' in header or 'name' in header or 'деталь' in header:
                                name_idx = i
                            elif 'ширина' in header or 'width' in header:
                                width_idx = i
                            elif 'высота' in header or 'height' in header:
                                height_idx = i
                            elif 'количество' in header or 'quantity' in header or 'кол-во' in header:
                                quantity_idx = i
                        
                        # Если нашли нужные колонки, парсим данные
                        if name_idx is not None and width_idx is not None and height_idx is not None:
                            for row in table[1:]:
                                try:
                                    if len(row) <= max(name_idx, width_idx, height_idx):
                                        continue
                                    
                                    name = str(row[name_idx]).strip() if row[name_idx] else f'Деталь из PDF'
                                    
                                    # Пытаемся извлечь числа из текста
                                    width_str = str(row[width_idx]) if row[width_idx] else '0'
                                    height_str = str(row[height_idx]) if row[height_idx] else '0'
                                    
                                    # Извлекаем числа (убираем текст)
                                    import re
                                    width_match = re.search(r'[\d.]+', width_str.replace(',', '.'))
                                    height_match = re.search(r'[\d.]+', height_str.replace(',', '.'))
                                    
                                    width = float(width_match.group()) if width_match else 0
                                    height = float(height_match.group()) if height_match else 0
                                    
                                    quantity = 1
                                    if quantity_idx is not None and row[quantity_idx]:
                                        qty_match = re.search(r'\d+', str(row[quantity_idx]))
                                        if qty_match:
                                            quantity = int(qty_match.group())
                                    
                                    if width > 0 and height > 0 and name:
                                        area_m2 = (width * height) / 1_000_000
                                        results.append({
                                            'name': name,
                                            'width': round(width, 1),
                                            'height': round(height, 1),
                                            'area_m2': round(area_m2, 4),
                                            'quantity': quantity
                                        })
                                except Exception as row_error:
                                    logger.warning(f"Ошибка обработки строки таблицы: {row_error}")
                                    continue
            
            # Удаляем временный файл
            try:
                filepath.unlink()
            except:
                pass
            
            if len(results) == 0:
                return jsonify({
                    'error': 'Не удалось извлечь данные из PDF. Убедитесь, что PDF содержит таблицу с колонками: Название, Ширина, Высота'
                }), 400
            
            logger.info(f"✓ Импортировано {len(results)} деталей из PDF")
            
            return jsonify({
                'success': True,
                'parts': results
            })
            
        except Exception as parse_error:
            logger.error(f"Ошибка парсинга PDF: {parse_error}", exc_info=True)
            return jsonify({'error': f'Ошибка чтения PDF: {str(parse_error)}'}), 500
        
    except Exception as e:
        logger.error(f"Ошибка импорта PDF: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/nesting/validate', methods=['POST'])
def validate_nesting():
    """
    Валидация раскроя - проверка координат и размещения деталей
    
    POST /api/nesting/validate
    {
        "sheets": [
            {
                "sheet_number": 1,
                "parts": [
                    {"name": "деталь", "x": 10, "y": 10, "width": 100, "height": 50, "rotated": false},
                    ...
                ]
            }
        ],
        "sheet_width": 2500,
        "sheet_height": 1250
    }
    
    Returns:
        {
            'valid': bool,
            'errors': [...],
            'warnings': [...],
            'details': {
                'total_parts': int,
                'intersections': [...],
                'out_of_bounds': [...],
                'overlaps': [...]
            }
        }
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Empty request body'}), 400
        
        sheets = data.get('sheets', [])
        sheet_width = data.get('sheet_width', 2500)
        sheet_height = data.get('sheet_height', 1250)
        
        errors = []
        warnings = []
        intersections = []
        out_of_bounds = []
        overlaps = []
        total_parts = 0
        
        logger.info(f"[VALIDATE] Валидация раскроя: {len(sheets)} листов, размер {sheet_width}x{sheet_height}")
        
        for sheet in sheets:
            sheet_num = sheet.get('sheet_number', 0)
            parts = sheet.get('parts', [])
            total_parts += len(parts)
            
            logger.info(f"[VALIDATE] Лист {sheet_num}: {len(parts)} деталей")
            
            # Проверяем каждую деталь
            for i, part in enumerate(parts):
                name = part.get('name', f'Деталь {i+1}')
                x = float(part.get('x', 0))
                y = float(part.get('y', 0))
                width = float(part.get('width', 0))
                height = float(part.get('height', 0))
                rotated = part.get('rotated', False)
                
                # Проверка 1: Выход за границы листа
                x2 = x + width
                y2 = y + height
                
                if x < 0 or y < 0:
                    error_msg = f"Лист {sheet_num}, {name}: координаты отрицательные (x={x:.1f}, y={y:.1f})"
                    errors.append(error_msg)
                    out_of_bounds.append({
                        'sheet': sheet_num,
                        'part': name,
                        'issue': 'negative_coords',
                        'x': x,
                        'y': y
                    })
                    logger.warning(f"[VALIDATE] {error_msg}")
                
                if x2 > sheet_width or y2 > sheet_height:
                    error_msg = f"Лист {sheet_num}, {name}: деталь выходит за границы листа (x2={x2:.1f} > {sheet_width} или y2={y2:.1f} > {sheet_height})"
                    errors.append(error_msg)
                    out_of_bounds.append({
                        'sheet': sheet_num,
                        'part': name,
                        'issue': 'out_of_bounds',
                        'x': x,
                        'y': y,
                        'x2': x2,
                        'y2': y2,
                        'sheet_width': sheet_width,
                        'sheet_height': sheet_height
                    })
                    logger.warning(f"[VALIDATE] {error_msg}")
                
                # Проверка 2: Некорректные размеры
                if width <= 0 or height <= 0:
                    error_msg = f"Лист {sheet_num}, {name}: некорректные размеры (width={width:.1f}, height={height:.1f})"
                    errors.append(error_msg)
                    logger.warning(f"[VALIDATE] {error_msg}")
                
                # Проверка 3: Пересечения с другими деталями на этом листе
                for j, other_part in enumerate(parts):
                    if i == j:
                        continue
                    
                    other_name = other_part.get('name', f'Деталь {j+1}')
                    other_x = float(other_part.get('x', 0))
                    other_y = float(other_part.get('y', 0))
                    other_width = float(other_part.get('width', 0))
                    other_height = float(other_part.get('height', 0))
                    other_x2 = other_x + other_width
                    other_y2 = other_y + other_height
                    
                    # Проверка пересечения прямоугольников
                    # Два прямоугольника НЕ пересекаются, если:
                    # - один полностью слева от другого: x2 <= other_x ИЛИ other_x2 <= x
                    # - один полностью сверху от другого: y2 <= other_y ИЛИ other_y2 <= y
                    # Если НЕ выполняется ни одно из этих условий, значит есть пересечение
                    if not (x2 <= other_x or other_x2 <= x or y2 <= other_y or other_y2 <= y):
                        # Есть пересечение
                        intersection_x_min = max(x, other_x)
                        intersection_x_max = min(x2, other_x2)
                        intersection_y_min = max(y, other_y)
                        intersection_y_max = min(y2, other_y2)
                        
                        intersection_width = intersection_x_max - intersection_x_min
                        intersection_height = intersection_y_max - intersection_y_min
                        intersection_area = intersection_width * intersection_height
                        
                        error_msg = f"Лист {sheet_num}: ПЕРЕСЕЧЕНИЕ '{name}' и '{other_name}' - площадь пересечения: {intersection_area:.1f} мм² ({intersection_width:.1f}×{intersection_height:.1f} мм)"
                        errors.append(error_msg)
                        
                        # Добавляем только если еще не добавлено (избегаем дубликатов)
                        already_added = False
                        for existing in intersections:
                            if (existing['part1'] == name and existing['part2'] == other_name) or \
                               (existing['part1'] == other_name and existing['part2'] == name):
                                already_added = True
                                break
                        
                        if not already_added:
                            intersections.append({
                                'sheet': sheet_num,
                                'part1': name,
                                'part2': other_name,
                                'part1_coords': {'x': x, 'y': y, 'x2': x2, 'y2': y2},
                                'part2_coords': {'x': other_x, 'y': other_y, 'x2': other_x2, 'y2': other_y2},
                                'intersection_area_mm2': intersection_area,
                                'intersection_width': intersection_width,
                                'intersection_height': intersection_height,
                                'intersection_coords': {
                                    'x': intersection_x_min,
                                    'y': intersection_y_min,
                                    'x2': intersection_x_max,
                                    'y2': intersection_y_max
                                }
                            })
                        logger.warning(f"[VALIDATE] {error_msg}")
                
                # Проверка 4: Слишком близкое расположение (меньше 5мм зазора)
                for j, other_part in enumerate(parts):
                    if i == j:
                        continue
                    
                    other_name = other_part.get('name', f'Деталь {j+1}')
                    other_x = float(other_part.get('x', 0))
                    other_y = float(other_part.get('y', 0))
                    other_width = float(other_part.get('width', 0))
                    other_height = float(other_part.get('height', 0))
                    other_x2 = other_x + other_width
                    other_y2 = other_y + other_height
                    
                    # Вычисляем минимальное расстояние между прямоугольниками
                    # Если они не пересекаются, вычисляем расстояние
                    if x2 <= other_x or other_x2 <= x or y2 <= other_y or other_y2 <= y:
                        # Не пересекаются, вычисляем расстояние
                        dx = max(0, max(x - other_x2, other_x - x2))
                        dy = max(0, max(y - other_y2, other_y - y2))
                        min_distance = max(dx, dy)
                        
                        if min_distance < 5.0 and min_distance > 0:
                            warning_msg = f"Лист {sheet_num}: {name} и {other_name} слишком близко (расстояние: {min_distance:.1f} мм < 5 мм)"
                            warnings.append(warning_msg)
                            overlaps.append({
                                'sheet': sheet_num,
                                'part1': name,
                                'part2': other_name,
                                'distance_mm': min_distance
                            })
                            logger.info(f"[VALIDATE] {warning_msg}")
        
        is_valid = len(errors) == 0
        
        result = {
            'valid': is_valid,
            'errors': errors,
            'warnings': warnings,
            'details': {
                'total_parts': total_parts,
                'total_sheets': len(sheets),
                'intersections': intersections,
                'out_of_bounds': out_of_bounds,
                'overlaps': overlaps,
                'errors_count': len(errors),
                'warnings_count': len(warnings)
            }
        }
        
        if is_valid:
            logger.info(f"[VALIDATE] ✓ Раскрой валиден: {total_parts} деталей на {len(sheets)} листах")
        else:
            logger.warning(f"[VALIDATE] ✗ Раскрой содержит ошибки: {len(errors)} ошибок, {len(warnings)} предупреждений")
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Ошибка валидации раскроя: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    """
    Экспорт результатов раскроя в Excel
    
    POST /api/export/excel
    {
        "nesting_result": {...},  # Результат раскроя
        "validation_result": {...}  # Результат валидации (опционально)
    }
    
    Returns: Excel файл
    """
    try:
        if not PANDAS_AVAILABLE:
            return jsonify({'error': 'pandas/openpyxl не установлены. Установите: pip install pandas openpyxl'}), 500
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Empty request body'}), 400
        
        nesting_result = data.get('nesting_result')
        if not nesting_result:
            return jsonify({'error': 'nesting_result is required'}), 400
        
        order_number = data.get('order_number', '')
        material_price = data.get('material_price', 0)
        
        # Создаем Excel файл
        import io
        from datetime import datetime
        
        output = io.BytesIO()
        
        try:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Лист 1: Сводная информация
                summary_params = ['Дата расчета', 'Размер листа (мм)', 'Количество листов',
                            'Использование материала (%)', 'Обрезки (%)',
                            'Площадь деталей (м²)', 'Площадь листов (м²)', 'Площадь обрезков (м²)']
                summary_values = [
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    f"{nesting_result.get('sheet_width', 2500)}×{nesting_result.get('sheet_height', 1250)}",
                    nesting_result.get('sheets_needed', 0),
                    f"{nesting_result.get('utilization_percent', 0)}%",
                    f"{nesting_result.get('waste_percent', 0)}%",
                    nesting_result.get('total_parts_area_m2', 0),
                    nesting_result.get('total_sheet_area_m2', 0),
                    nesting_result.get('total_waste_area_m2', 0)
                ]
                
                # Добавляем номер заказа, если указан
                if order_number:
                    summary_params.insert(0, 'Номер заказа')
                    summary_values.insert(0, order_number)
                
                # Добавляем информацию о материале, если указана цена
                material_name = data.get('material_name', '')
                if material_name:
                    summary_params.insert(1, 'Материал')
                    summary_values.insert(1, material_name)
                
                # Добавляем финансовый отчет, если указана цена
                if material_price > 0:
                    total_parts_area = nesting_result.get('total_parts_area_m2', 0)
                    total_sheet_area = nesting_result.get('total_sheet_area_m2', 0)
                    total_waste_area = nesting_result.get('total_waste_area_m2', 0)
                    
                    parts_cost = total_parts_area * material_price
                    waste_cost = total_waste_area * material_price
                    total_cost = total_sheet_area * material_price
                    
                    summary_params.extend(['', 'ФИНАНСОВЫЙ ОТЧЕТ', 'Цена материала (₽/м²)',
                                          'Стоимость деталей (₽)', 'Стоимость обрезков (₽)',
                                          'Общая стоимость (₽)', 'Эффективность использования (%)'])
                    summary_values.extend(['', '', material_price, parts_cost, waste_cost,
                                         total_cost, f"{((parts_cost / total_cost) * 100) if total_cost > 0 else 0:.1f}%"])
                
                summary_data = {
                    'Параметр': summary_params,
                    'Значение': summary_values
                }
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Сводка', index=False, startrow=1)
                
                # Добавляем заголовок таблицы и автоширину колонок
                worksheet = writer.sheets['Сводка']
                try:
                    from openpyxl.styles import Font, Alignment
                    worksheet.merge_cells('A1:B1')
                    header_cell = worksheet['A1']
                    header_cell.value = 'СВОДНАЯ ИНФОРМАЦИЯ'
                    header_cell.font = Font(bold=True, size=14)
                    header_cell.alignment = Alignment(horizontal='center', vertical='center')
                except Exception as header_error:
                    logger.warning(f"Не удалось добавить заголовок для листа 'Сводка': {header_error}")
                
                # Автоширина колонок и выравнивание для листа "Сводка"
                try:
                    from openpyxl.cell.cell import MergedCell
                    from openpyxl.styles import Alignment
                    # Выравнивание заголовков колонок (row 2) - по центру
                    header_row = worksheet[2]
                    if header_row[0]:  # Колонка A (Параметр)
                        header_row[0].alignment = Alignment(horizontal='center', vertical='center')
                    if header_row[1]:  # Колонка B (Значение)
                        header_row[1].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Выравнивание данных: колонка A (Параметр) - по левому краю, колонка B (Значение) - по центру
                    for row_idx, row in enumerate(worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=2), start=3):
                        # Колонка A - по левому краю
                        if row[0]:
                            row[0].alignment = Alignment(horizontal='left', vertical='center')
                        # Колонка B - по центру (значения обычно числовые или процентные)
                        if row[1]:
                            row[1].alignment = Alignment(horizontal='center', vertical='center')
                    
                    for column in worksheet.columns:
                        max_length = 0
                        # Пропускаем объединенные ячейки
                        first_cell = None
                        for cell in column:
                            if isinstance(cell, MergedCell):
                                continue
                            if first_cell is None:
                                first_cell = cell
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        if first_cell:
                            column_letter = first_cell.column_letter
                            adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                except Exception as width_error:
                    logger.warning(f"Не удалось установить автоширину для листа 'Сводка': {width_error}")
                
                # Лист 2: Таблица позиций
                if nesting_result.get('positions_summary'):
                    positions_data = []
                    for pos in nesting_result['positions_summary']:
                        pos_dict = {
                            '№ поз.': pos.get('position_number', ''),
                            'Наименование': pos.get('name', ''),
                            'Ширина (мм)': pos.get('width', 0),
                            'Высота (мм)': pos.get('height', 0),
                            'Площадь 1шт (м²)': pos.get('area_m2', 0),
                            'Количество': pos.get('quantity', 0),
                            'Итого площадь (м²)': pos.get('total_area_m2', 0)
                        }
                        # Добавляем стоимость, если указана цена
                        if material_price > 0:
                            pos_dict['Стоимость 1шт (₽)'] = round(pos.get('area_m2', 0) * material_price, 2)
                            pos_dict['Итого стоимость (₽)'] = round(pos.get('total_area_m2', 0) * material_price, 2)
                        positions_data.append(pos_dict)
                    
                    df_positions = pd.DataFrame(positions_data)
                    df_positions.to_excel(writer, sheet_name='Позиции', index=False, startrow=1)
                    
                    # Добавляем заголовок таблицы и автоширину колонок
                    worksheet = writer.sheets['Позиции']
                    try:
                        from openpyxl.styles import Font, Alignment
                        # Объединяем ячейки для заголовка
                        num_cols = len(df_positions.columns)
                        last_col_letter = worksheet.cell(row=1, column=num_cols).column_letter
                        worksheet.merge_cells(f'A1:{last_col_letter}1')
                        header_cell = worksheet['A1']
                        header_cell.value = 'ТАБЛИЦА ПОЗИЦИЙ'
                        header_cell.font = Font(bold=True, size=14)
                        header_cell.alignment = Alignment(horizontal='center', vertical='center')
                    except Exception as header_error:
                        logger.warning(f"Не удалось добавить заголовок для листа 'Позиции': {header_error}")
                    
                    # Автоширина колонок и выравнивание для листа "Позиции"
                    try:
                        from openpyxl.cell.cell import MergedCell
                        from openpyxl.styles import Alignment
                        # Выравнивание заголовков колонок (row 2) - по центру
                        header_row = worksheet[2]
                        for col_idx, cell in enumerate(header_row, start=1):
                            if cell:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Выравнивание данных: колонка "Наименование" (B) - по левому краю, остальные - по центру
                        for row_idx, row in enumerate(worksheet.iter_rows(min_row=3, max_row=worksheet.max_row), start=3):
                            for col_idx, cell in enumerate(row, start=1):
                                if cell:
                                    # Колонка B (Наименование) - по левому краю, остальные - по центру
                                    if col_idx == 2:  # Наименование
                                        cell.alignment = Alignment(horizontal='left', vertical='center')
                                    else:
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        for column in worksheet.columns:
                            max_length = 0
                            # Пропускаем объединенные ячейки
                            first_cell = None
                            for cell in column:
                                if isinstance(cell, MergedCell):
                                    continue
                                if first_cell is None:
                                    first_cell = cell
                                try:
                                    if cell.value and len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            if first_cell:
                                column_letter = first_cell.column_letter
                                adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
                                worksheet.column_dimensions[column_letter].width = adjusted_width
                    except Exception as width_error:
                        logger.warning(f"Не удалось установить автоширину для листа 'Позиции': {width_error}")
            
                # Лист 3+: Координаты для каждого листа
                for sheet in nesting_result.get('sheets', []):
                    sheet_num = sheet.get('sheet_number', 1)
                    parts_data = []
                    for part in sheet.get('parts', []):
                        part_dict = {
                            '№ поз.': part.get('position_number', ''),
                            'Наименование': part.get('name', ''),
                            'X (мм)': round(part.get('x', 0), 1),
                            'Y (мм)': round(part.get('y', 0), 1),
                            'Ширина (мм)': round(part.get('width', 0), 1),
                            'Высота (мм)': round(part.get('height', 0), 1),
                            'X2 (мм)': round(part.get('x', 0) + part.get('width', 0), 1),
                            'Y2 (мм)': round(part.get('y', 0) + part.get('height', 0), 1),
                            'Поворот': 'Да' if part.get('rotated', False) else 'Нет',
                            'Площадь (м²)': round(part.get('area_m2', 0), 4)
                        }
                        # Добавляем стоимость, если указана цена
                        if material_price > 0:
                            part_dict['Стоимость (₽)'] = round(part.get('area_m2', 0) * material_price, 2)
                        parts_data.append(part_dict)
                    df_sheet = pd.DataFrame(parts_data)
                    # Добавляем заголовок с номером заказа, если указан
                    sheet_name = f'Лист {sheet_num}'
                    if order_number:
                        sheet_name = f'Лист {sheet_num} ({order_number})'
                    
                    df_sheet.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                    
                    # Добавляем заголовок таблицы и автоширину колонок
                    worksheet = writer.sheets[sheet_name]
                    try:
                        from openpyxl.styles import Font, Alignment
                        # Объединяем ячейки для заголовка
                        num_cols = len(df_sheet.columns)
                        last_col_letter = worksheet.cell(row=1, column=num_cols).column_letter
                        worksheet.merge_cells(f'A1:{last_col_letter}1')
                        header_cell = worksheet['A1']
                        header_cell.value = f'КООРДИНАТЫ РАЗМЕЩЕНИЯ ДЕТАЛЕЙ - ЛИСТ {sheet_num}'
                        header_cell.font = Font(bold=True, size=14)
                        header_cell.alignment = Alignment(horizontal='center', vertical='center')
                    except Exception as header_error:
                        logger.warning(f"Не удалось добавить заголовок для листа '{sheet_name}': {header_error}")
                    
                    # Автоширина колонок и выравнивание для каждого листа
                    try:
                        from openpyxl.cell.cell import MergedCell
                        from openpyxl.styles import Alignment
                        # Выравнивание заголовков колонок (row 2) - по центру
                        header_row = worksheet[2]
                        for col_idx, cell in enumerate(header_row, start=1):
                            if cell:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Выравнивание данных: колонка "Наименование" (B) - по левому краю, остальные - по центру
                        for row_idx, row in enumerate(worksheet.iter_rows(min_row=3, max_row=worksheet.max_row), start=3):
                            for col_idx, cell in enumerate(row, start=1):
                                if cell:
                                    # Колонка B (Наименование) - по левому краю, остальные - по центру
                                    if col_idx == 2:  # Наименование
                                        cell.alignment = Alignment(horizontal='left', vertical='center')
                                    else:
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        for column in worksheet.columns:
                            max_length = 0
                            # Пропускаем объединенные ячейки
                            first_cell = None
                            for cell in column:
                                if isinstance(cell, MergedCell):
                                    continue
                                if first_cell is None:
                                    first_cell = cell
                                try:
                                    if cell.value and len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            if first_cell:
                                column_letter = first_cell.column_letter
                                adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
                                worksheet.column_dimensions[column_letter].width = adjusted_width
                    except Exception as width_error:
                        logger.warning(f"Не удалось установить автоширину для листа '{sheet_name}': {width_error}")
        except Exception as excel_error:
            logger.error(f"Ошибка при создании Excel файла: {excel_error}", exc_info=True)
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            return jsonify({'error': f'Ошибка создания Excel файла: {str(excel_error)}'}), 500
        
        # Формируем имя файла с номером заказа
        try:
            if order_number:
                # Убираем недопустимые символы для имени файла
                safe_order_number = order_number.replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_')
                filename = f"Расчет площади и раскроя {safe_order_number}.xlsx"
            else:
                filename = f"Расчет площади и раскроя {datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            logger.info(f"✓ Экспортирован Excel файл: {filename}")
            
            # Перемещаем указатель в начало файла и читаем данные
            file_data = None
            file_size = 0
            try:
                output.seek(0)
                file_data = output.getvalue()
                file_size = len(file_data)
                logger.info(f"Размер файла: {file_size} байт")
                
                if file_size == 0:
                    logger.error("Файл пуст после записи!")
                    return jsonify({'error': 'Ошибка создания Excel файла: файл пуст'}), 500
            except Exception as seek_error:
                logger.error(f"Ошибка при чтении файла: {seek_error}", exc_info=True)
                import traceback
                logger.error(f"Traceback: {traceback.format_exc()}")
                return jsonify({'error': f'Ошибка чтения Excel файла: {str(seek_error)}'}), 500
            
            if file_data is None:
                logger.error("file_data is None после чтения!")
                return jsonify({'error': 'Ошибка: не удалось прочитать данные файла'}), 500
            
            import urllib.parse
            # Используем безопасное имя файла для заголовка
            safe_filename = filename.encode('utf-8')
            encoded_filename = urllib.parse.quote(safe_filename)
            
            logger.info(f"Отправка файла: {filename}, размер: {file_size} байт")
            # Создаем новый BytesIO из данных, так как output может быть закрыт
            output_copy = io.BytesIO(file_data)
            output_copy.seek(0)
            response = send_file(
                output_copy,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            # Устанавливаем заголовок для правильного имени файла
            # Используем только filename* для избежания проблем с кодировкой
            response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
            logger.info(f"Файл успешно подготовлен для отправки")
            return response
        except Exception as send_error:
            logger.error(f"Ошибка при отправке Excel файла: {send_error}", exc_info=True)
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            return jsonify({'error': f'Ошибка отправки Excel файла: {str(send_error)}'}), 500
        
    except Exception as e:
        logger.error(f"Ошибка экспорта Excel: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/pdf', methods=['POST'])
def export_pdf():
    """
    Экспорт результатов раскроя в PDF
    
    POST /api/export/pdf
    {
        "nesting_result": {...},  # Результат раскроя
        "validation_result": {...}  # Результат валидации (опционально)
    }
    
    Returns: PDF файл
    """
    try:
        if not REPORTLAB_AVAILABLE:
            return jsonify({'error': 'reportlab не установлен. Установите: pip install reportlab'}), 500
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Empty request body'}), 400
        
        nesting_result = data.get('nesting_result')
        if not nesting_result:
            return jsonify({'error': 'nesting_result is required'}), 400
        
        validation_result = data.get('validation_result')
        material_price = data.get('material_price', 0)  # Цена материала за м²
        material_name = data.get('material_name', '')  # Название материала
        order_number = data.get('order_number', '')
        
        # Создаем PDF
        import io
        from datetime import datetime
        
        buffer = io.BytesIO()
        # Отключаем колонтитулы для чистого PDF, используем альбомную ориентацию
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=landscape(A4),
            leftMargin=15*mm,
            rightMargin=15*mm,
            topMargin=15*mm,
            bottomMargin=15*mm,
            title="Результаты раскроя",
            author="ZVD Calculator"
        )
        story = []
        styles = getSampleStyleSheet()
        
        # Регистрируем шрифт с поддержкой кириллицы
        # Используем системный шрифт Windows с поддержкой русского языка
        font_found = False
        font_name = 'Helvetica'
        bold_font = 'Helvetica-Bold'
        
        try:
            from reportlab.pdfbase.ttfonts import TTFont
            import os
            # Ищем шрифт в системных папках Windows
            font_paths = [
                ('C:/Windows/Fonts/arial.ttf', 'C:/Windows/Fonts/arialbd.ttf'),
                ('C:/Windows/Fonts/calibri.ttf', 'C:/Windows/Fonts/calibrib.ttf'),
                ('C:/Windows/Fonts/times.ttf', 'C:/Windows/Fonts/timesbd.ttf'),
            ]
            for normal_path, bold_path in font_paths:
                if os.path.exists(normal_path):
                    try:
                        # Регистрируем обычный шрифт
                        pdfmetrics.registerFont(TTFont('RussianFont', normal_path, subfontIndex=0))
                        font_found = True
                        font_name = 'RussianFont'
                        
                        # Пробуем зарегистрировать жирный вариант
                        if os.path.exists(bold_path):
                            try:
                                pdfmetrics.registerFont(TTFont('RussianFont-Bold', bold_path, subfontIndex=0))
                                bold_font = 'RussianFont-Bold'
                            except:
                                bold_font = 'RussianFont'
                        else:
                            bold_font = 'RussianFont'
                        logger.info(f"Зарегистрирован шрифт: {normal_path}, жирный: {bold_font}")
                        break
                    except Exception as e:
                        logger.warning(f"Не удалось зарегистрировать шрифт {font_path}: {e}")
                        continue
            
            if not font_found:
                logger.warning("Не найден TTF шрифт с кириллицей, используем стандартный (могут быть проблемы с русским текстом)")
        except Exception as font_error:
            logger.warning(f"Ошибка регистрации шрифта: {font_error}, используем стандартный")
        
        # Заголовок
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1976D2'),
            spaceAfter=30,
            fontName=font_name
        )
        # Формируем заголовок с номером заказа, если указан
        title_text = 'Результаты раскроя'
        if order_number:
            title_text = f'Результаты раскроя {order_number}'
        story.append(Paragraph(title_text, title_style))
        story.append(Spacer(1, 12))
        
        # Сводная информация
        heading2_style = ParagraphStyle(
            'CustomHeading2',
            parent=styles['Heading2'],
            fontName=font_name
        )
        story.append(Paragraph('<b>Сводная информация</b>', heading2_style))
        
        # Рассчитываем финансовые показатели
        total_parts_area = nesting_result.get('total_parts_area_m2', 0)
        total_sheet_area = nesting_result.get('total_sheet_area_m2', 0)
        total_waste_area = nesting_result.get('total_waste_area_m2', 0)
        
        total_cost = total_sheet_area * material_price if material_price > 0 else 0
        parts_cost = total_parts_area * material_price if material_price > 0 else 0
        waste_cost = total_waste_area * material_price if material_price > 0 else 0
        
        summary_data = [
            ['Параметр', 'Значение']
        ]
        
        # Добавляем номер заказа, если указан
        if order_number:
            summary_data.append(['Номер заказа', order_number])
        
        # Добавляем информацию о материале, если указана
        if material_name:
            summary_data.append(['Материал', material_name])
        
        # Добавляем остальные параметры
        summary_data.extend([
            ['Дата расчета', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Размер листа', f"{nesting_result.get('sheet_width', 2500)}×{nesting_result.get('sheet_height', 1250)} мм"],
            ['Количество листов', str(nesting_result.get('sheets_needed', 0))],
            ['Использование материала', f"{nesting_result.get('utilization_percent', 0)}%"],
            ['Обрезки', f"{nesting_result.get('waste_percent', 0)}%"],
            ['Площадь деталей', f"{total_parts_area:.4f} м²"],
            ['Площадь листов', f"{total_sheet_area:.4f} м²"],
            ['Площадь обрезков', f"{total_waste_area:.4f} м²"]
        ])
        
        # Добавляем финансовый отчет, если указана цена
        if material_price > 0:
            summary_data.extend([
                ['', ''],
                ['<b>ФИНАНСОВЫЙ ОТЧЕТ</b>', ''],
                ['Цена материала', f"{material_price:.2f} ₽/м²"],
                ['Стоимость деталей', f"{parts_cost:.2f} ₽"],
                ['Стоимость обрезков', f"{waste_cost:.2f} ₽"],
                ['<b>Общая стоимость</b>', f"<b>{total_cost:.2f} ₽</b>"],
                ['Эффективность использования', f"{((parts_cost / total_cost) * 100) if total_cost > 0 else 0:.1f}%"]
            ])
        
        # Используем Paragraph для всех ячеек с правильным шрифтом
        summary_data_with_font = []
        header_style = ParagraphStyle('SummaryHeader', fontName=bold_font, fontSize=13)
        data_style = ParagraphStyle('SummaryData', fontName=font_name, fontSize=11)
        bold_data_style = ParagraphStyle('SummaryDataBold', fontName=bold_font, fontSize=11)
        
        for row_idx, row in enumerate(summary_data):
            if row_idx == 0:
                # Заголовки
                header_row = [
                    Paragraph(row[0], header_style),
                    Paragraph(row[1], header_style)
                ]
                summary_data_with_font.append(header_row)
            else:
                # Данные - проверяем на жирный текст
                cell0 = str(row[0])
                cell1 = str(row[1])
                
                # Если содержит <b> теги, используем жирный шрифт
                if '<b>' in cell0 or '<b>' in cell1:
                    # Убираем HTML теги и используем жирный стиль
                    cell0_clean = cell0.replace('<b>', '').replace('</b>', '')
                    cell1_clean = cell1.replace('<b>', '').replace('</b>', '')
                    data_row = [
                        Paragraph(cell0_clean, bold_data_style),
                        Paragraph(cell1_clean, bold_data_style)
                    ]
                else:
                    data_row = [
                        Paragraph(cell0, data_style),
                        Paragraph(cell1, data_style)
                    ]
                summary_data_with_font.append(data_row)
        
        summary_table = Table(summary_data_with_font, colWidths=[100*mm, 80*mm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
            ('TOPPADDING', (0, 0), (-1, 0), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black)
        ]))
        story.append(summary_table)
        story.append(PageBreak())  # Новая страница для таблицы позиций
        
        # Таблица позиций
        if nesting_result.get('positions_summary'):
            positions_heading_style = ParagraphStyle(
                'PositionsHeading',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1976D2'),
                spaceAfter=20,
                fontName=font_name
            )
            story.append(Paragraph('<b>Таблица позиций</b>', positions_heading_style))
            
            positions_data = [['№ поз.', 'Наименование', 'Ширина (мм)', 'Высота (мм)', 'Площадь 1шт (м²)', 'Количество', 'Итого площадь (м²)']]
            if material_price > 0:
                positions_data[0].extend(['Стоимость 1шт (₽)', 'Итого стоимость (₽)'])
            
            for pos in nesting_result['positions_summary']:
                row = [
                    str(pos.get('position_number', '')),
                    pos.get('name', ''),
                    str(pos.get('width', 0)),
                    str(pos.get('height', 0)),
                    f"{pos.get('area_m2', 0):.4f}",
                    str(pos.get('quantity', 0)),
                    f"{pos.get('total_area_m2', 0):.4f}"
                ]
                if material_price > 0:
                    row.append(f"{pos.get('area_m2', 0) * material_price:.2f}")
                    row.append(f"{pos.get('total_area_m2', 0) * material_price:.2f}")
                positions_data.append(row)
            
            # Используем Paragraph для всех ячеек
            positions_data_with_font = []
            pos_header_style = ParagraphStyle('PosHeader', fontName=bold_font, fontSize=10)
            pos_data_style = ParagraphStyle('PosData', fontName=font_name, fontSize=9)
            
            for row_idx, row in enumerate(positions_data):
                if row_idx == 0:
                    header_row = [Paragraph(str(cell), pos_header_style) for cell in row]
                    positions_data_with_font.append(header_row)
                else:
                    data_row = []
                    for col_idx, cell in enumerate(row):
                        if col_idx == 1:  # Колонка "Наименование" - делаем перенос текста
                            # Создаем стиль с переносом слов
                            name_style = ParagraphStyle(
                                'PosNameData',
                                fontName=font_name,
                                fontSize=8,  # Уменьшаем шрифт для длинных названий
                                leading=9,
                                wordWrap='CJK'  # Перенос слов
                            )
                            # Ограничиваем длину названия и добавляем перенос
                            cell_text = str(cell)
                            if len(cell_text) > 40:
                                # Разбиваем длинные названия на части
                                words = cell_text.split()
                                lines = []
                                current_line = ''
                                for word in words:
                                    if len(current_line + word) < 40:
                                        current_line += word + ' '
                                    else:
                                        if current_line:
                                            lines.append(current_line.strip())
                                        current_line = word + ' '
                                if current_line:
                                    lines.append(current_line.strip())
                                cell_text = '<br/>'.join(lines[:3])  # Максимум 3 строки
                            data_row.append(Paragraph(cell_text, name_style))
                        else:
                            data_row.append(Paragraph(str(cell), pos_data_style))
                    positions_data_with_font.append(data_row)
            
            # Ширина колонок для альбомной ориентации (больше места)
            col_widths = [20*mm, 70*mm, 25*mm, 25*mm, 28*mm, 20*mm, 28*mm]
            if material_price > 0:
                col_widths.extend([28*mm, 28*mm])
            
            positions_table = Table(positions_data_with_font, colWidths=col_widths, repeatRows=1)
            positions_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Выравнивание по верху для многострочных названий
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
                ('TOPPADDING', (0, 0), (-1, 0), 5),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 2),
                ('TOPPADDING', (0, 1), (-1, -1), 2),
                ('LEFTPADDING', (0, 0), (-1, -1), 2),
                ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black)
            ]))
            story.append(positions_table)
            story.append(PageBreak())  # Новая страница для визуализации раскроя
        
        # Результаты валидации
        if validation_result:
            validation_style = ParagraphStyle(
                'CustomHeading2',
                parent=styles['Heading2'],
                fontName=font_name
            )
            story.append(Paragraph('<b>Результаты проверки</b>', validation_style))
            normal_style = ParagraphStyle('Normal', fontName=font_name, fontSize=10)
            if validation_result.get('valid'):
                story.append(Paragraph('✓ Раскрой валиден', normal_style))
            else:
                story.append(Paragraph('✗ Обнаружены ошибки', normal_style))
                for error in validation_result.get('errors', [])[:5]:  # Первые 5 ошибок
                    story.append(Paragraph(f'• {error}', normal_style))
            story.append(Spacer(1, 20))
        
        # Визуализация и координаты для каждого листа
        for sheet in nesting_result.get('sheets', []):
            sheet_num = sheet.get('sheet_number', 1)
            sheet_width = nesting_result.get('sheet_width', 2500)
            sheet_height = nesting_result.get('sheet_height', 1250)
            
            # Заголовок листа
            sheet_heading_style = ParagraphStyle(
                'SheetHeading',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1976D2'),
                spaceAfter=15,
                fontName=font_name
            )
            story.append(Paragraph(f'<b>Лист {sheet_num}</b>', sheet_heading_style))
            
            # Визуализация раскроя (горизонтальная ориентация)
            try:
                # Для альбомной ориентации A4: ширина ~270mm, высота ~180mm
                # Лист 1250x2500 (ширина x длина) - нужно расположить горизонтально
                # В PDF: ширина листа (1250) идет по горизонтали, длина (2500) по вертикали
                # Но для визуализации нужно учесть, что лист должен быть горизонтальным
                # Поэтому: vis_width соответствует sheet_width (1250), vis_height соответствует sheet_height (2500)
                max_width = 250*mm  # Максимальная ширина для landscape A4
                scale = max_width / sheet_width  # Масштаб по ширине
                vis_height = sheet_height * scale
                
                # Ограничиваем высоту визуализации
                if vis_height > 150*mm:
                    scale = 150*mm / sheet_height
                    vis_width = sheet_width * scale
                    vis_height = 150*mm
                else:
                    vis_width = max_width
                
                drawing = Drawing(vis_width, vis_height)
                
                # Фон листа
                sheet_rect = Rect(0, 0, vis_width, vis_height, 
                                fillColor=colors.lightgrey, 
                                strokeColor=colors.black,
                                strokeWidth=1)
                drawing.add(sheet_rect)
                
                # Цвета для деталей
                part_colors = [
                    colors.HexColor('#4CAF50'), colors.HexColor('#2196F3'), 
                    colors.HexColor('#FF9800'), colors.HexColor('#9C27B0'),
                    colors.HexColor('#F44336'), colors.HexColor('#00BCD4'),
                    colors.HexColor('#FFC107'), colors.HexColor('#E91E63')
                ]
                
                # Рисуем детали
                for idx, part in enumerate(sheet.get('parts', [])):
                    x = part.get('x', 0) * scale
                    y = vis_height - (part.get('y', 0) + part.get('height', 0)) * scale  # Инвертируем Y
                    width = part.get('width', 0) * scale
                    height = part.get('height', 0) * scale
                    
                    color = part_colors[idx % len(part_colors)]
                    
                    # Проверяем пересечения
                    is_intersecting = False
                    if validation_result and not validation_result.get('valid'):
                        for intersection in validation_result.get('details', {}).get('intersections', []):
                            if (intersection.get('part1') == part.get('name') or 
                                intersection.get('part2') == part.get('name')):
                                is_intersecting = True
                                break
                    
                    part_rect = Rect(x, y, width, height,
                                   fillColor=colors.HexColor('#FF0000') if is_intersecting else color,
                                   strokeColor=colors.black,
                                   strokeWidth=2 if is_intersecting else 1,
                                   fillOpacity=0.7)
                    drawing.add(part_rect)
                    
                    # Номер позиции - временно отключен, так как String может вызывать ошибки
                    # Можно добавить позже через другой метод или использовать Image
                
                story.append(drawing)
                story.append(Spacer(1, 10))
                
                # Информация о листе
                info_style = ParagraphStyle('SheetInfo', fontName=font_name, fontSize=9)
                info_text = f"Использовано: {sheet.get('used_area_m2', 0):.4f} м² ({sheet.get('utilization_percent', 0)}%) | " \
                           f"Остаток: {sheet.get('waste_area_m2', 0):.4f} м²"
                story.append(Paragraph(info_text, info_style))
                story.append(Spacer(1, 15))
            except Exception as vis_error:
                logger.warning(f"Не удалось создать визуализацию для листа {sheet_num}: {vis_error}", exc_info=True)
            
            # Заголовок таблицы координат
            coord_heading_style = ParagraphStyle(
                'CoordHeading',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=10,
                fontName=font_name
            )
            story.append(Paragraph('<b>Координаты размещения деталей</b>', coord_heading_style))
            
            # Таблица координат
            coord_data = [['№ поз.', 'Наименование', 'X (мм)', 'Y (мм)', 'Ширина (мм)', 'Высота (мм)', 'X2 (мм)', 'Y2 (мм)', 'Поворот', 'Площадь (м²)']]
            if material_price > 0:
                coord_data[0].append('Стоимость (₽)')
            
            for part in sheet.get('parts', []):
                row = [
                    str(part.get('position_number', '')),
                    part.get('name', ''),
                    f"{part.get('x', 0):.1f}",
                    f"{part.get('y', 0):.1f}",
                    f"{part.get('width', 0):.1f}",
                    f"{part.get('height', 0):.1f}",
                    f"{part.get('x', 0) + part.get('width', 0):.1f}",
                    f"{part.get('y', 0) + part.get('height', 0):.1f}",
                    'Да' if part.get('rotated', False) else 'Нет',
                    f"{part.get('area_m2', 0):.4f}"
                ]
                if material_price > 0:
                    row.append(f"{part.get('area_m2', 0) * material_price:.2f}")
                coord_data.append(row)
            
            # Используем Paragraph для всех ячеек с правильным шрифтом
            coord_data_with_font = []
            # Заголовки с правильным шрифтом
            header_style = ParagraphStyle('TableHeader', fontName=bold_font, fontSize=11)
            header_row = []
            for header in coord_data[0]:
                # Используем простой текст - шрифт уже установлен в стиле
                header_row.append(Paragraph(header, header_style))
            coord_data_with_font.append(header_row)
            
            # Данные
            normal_style = ParagraphStyle('TableData', fontName=font_name, fontSize=9)
            name_style = ParagraphStyle(
                'CoordNameData',
                fontName=font_name,
                fontSize=8,  # Уменьшаем шрифт для названий
                leading=9,
                wordWrap='CJK'  # Перенос слов
            )
            for row in coord_data[1:]:
                data_row = []
                for col_idx, cell in enumerate(row):
                    if col_idx == 1:  # Колонка "Наименование" - делаем перенос текста
                        cell_text = str(cell)
                        if len(cell_text) > 30:
                            # Разбиваем длинные названия на части
                            words = cell_text.split()
                            lines = []
                            current_line = ''
                            for word in words:
                                if len(current_line + word) < 30:
                                    current_line += word + ' '
                                else:
                                    if current_line:
                                        lines.append(current_line.strip())
                                    current_line = word + ' '
                            if current_line:
                                lines.append(current_line.strip())
                            cell_text = '<br/>'.join(lines[:2])  # Максимум 2 строки
                        data_row.append(Paragraph(cell_text, name_style))
                    else:
                        data_row.append(Paragraph(str(cell), normal_style))
                coord_data_with_font.append(data_row)
            
            # Уменьшаем ширину колонок для компактности
            # Базовые колонки: № поз., Наименование, X, Y, Ширина, Высота, X2, Y2, Поворот, Площадь
            col_widths = [15*mm, 45*mm, 16*mm, 16*mm, 20*mm, 20*mm, 16*mm, 16*mm, 15*mm, 20*mm]
            if material_price > 0:
                col_widths.append(20*mm)  # Стоимость
            coord_table = Table(coord_data_with_font, colWidths=col_widths, repeatRows=1)
            coord_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Выравнивание по верху для многострочных названий
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
                ('TOPPADDING', (0, 0), (-1, 0), 5),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 2),
                ('TOPPADDING', (0, 1), (-1, -1), 2),
                ('LEFTPADDING', (0, 0), (-1, -1), 2),
                ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black)
            ]))
            story.append(coord_table)
            
            # Разрыв страницы после каждого листа (кроме последнего)
            if sheet_num < len(nesting_result.get('sheets', [])):
                story.append(PageBreak())
        
        doc.build(story)
        buffer.seek(0)
        
        # Формируем имя файла с номером заказа (как в Excel)
        if order_number:
            # Убираем недопустимые символы для имени файла
            safe_order_number = order_number.replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_')
            filename = f"Расчет площади и раскроя {safe_order_number}.pdf"
        else:
            filename = f"Расчет площади и раскроя {datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        logger.info(f"✓ Экспортирован PDF файл: {filename}")
        
        import urllib.parse
        # Используем безопасное имя файла для заголовка
        safe_filename = filename.encode('utf-8')
        encoded_filename = urllib.parse.quote(safe_filename)
        
        response = send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        # Устанавливаем заголовок для правильного имени файла
        # Используем только filename* для избежания проблем с кодировкой
        response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
        return response
        
    except Exception as e:
        logger.error(f"Ошибка экспорта PDF: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    logger.info("🚀 Запуск ZVD Area Calculator")
    logger.info("📡 API: http://localhost:5000")
    
    app.run(host='0.0.0.0', port=5000, debug=True)
